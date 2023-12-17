import pyodbc
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv
import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill
import subprocess
import glob


load_dotenv("../.env")

server = os.getenv("server")
database = os.getenv("database")
conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes;'
border = Border(left=Side(style='thin', color='6D9712'),
                right=Side(style='thin', color='6D9712'),
                top=Side(style='thin', color='6D9712'),
                bottom=Side(style='thin', color='6D9712'))
background_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')


def generate_file_recipes():
    files = glob.glob('./xlsx_files/*')
    for f in files:
        os.remove(f)
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    sender_files = {}
    try:
        # DietForms
        query = f"SELECT * FROM DietForms WHERE GenerateFile = 1"
        cursor.execute(query)
        diet_forms = cursor.fetchall()

        for diet_form in diet_forms:
            query_update = "UPDATE DietForms SET GenerateFile = 0 WHERE Id = ?"
            cursor.execute(query_update, diet_form[0])
            conn.commit()

        for diet_form in diet_forms:
            query = f"SELECT * FROM Users WHERE Id = ? "
            cursor.execute(query, diet_form[1])
            users = cursor.fetchall()
            file_path = f'./xlsx_files/{diet_form[0]}_recipes.xlsx'
            sender_files[users[0][3]] = file_path

            query = f"SELECT * FROM Recipes WHERE DietFormId = ? "
            cursor.execute(query, diet_form[0])
            recipes = cursor.fetchall()

            workbook = op.Workbook()
            days_of_week = {
                'Monday': 1,
                'Tuesday': 2,
                'Wednesday': 3,
                'Thursday': 4,
                'Friday': 5,
                'Saturday': 6,
                'Sunday': 7
            }
            row_meals = ['Breakfast', 'Lunch', 'Dinner']
            for key in days_of_week:
                workbook.create_sheet(title=key)
                start_x = 2
                ws = workbook[key]
                ws.sheet_view.showGridLines = False
                recipes_arr_daily = []
                for row_meal in row_meals:
                    for item in recipes:
                        if item[1] == days_of_week[key] and item[2] == row_meal:
                            recipes_arr_daily.append(item)

                for item in recipes_arr_daily:
                    start_y = 2
                    ws.column_dimensions[get_column_letter(start_x)].width = 60
                    ws.cell(row=start_y, column=start_x).value = f"{item[2]} -- {item[3]}"
                    ws.cell(row=start_y, column=start_x).font = op.styles.Font(bold=True)
                    start_y += 1
                    ws.cell(row=start_y, column=start_x).value = f"{item[4]}"
                    start_y += 1
                    ws.cell(row=start_y, column=start_x).value = f"{', '.join(item[10].split('////'))}"
                    ws.cell(row=start_y, column=start_x).alignment = op.styles.Alignment(wrap_text=True,
                                                                                         vertical='top')
                    ws.cell(row=start_y, column=start_x).border = border
                    ws.cell(row=start_y, column=start_x).fill = background_fill

                    ws.row_dimensions[start_y].auto_size = True
                    start_y += 4

                    query = f"SELECT * FROM RecipeInstructions WHERE RecipeId = ? ORDER BY [RecipeInstructions].[Order]"
                    cursor.execute(query, item[0])
                    recipes_instructions = cursor.fetchall()
                    ws.cell(row=start_y, column=start_x).value = "Instructions: "
                    ws.cell(row=start_y, column=start_x).font = op.styles.Font(bold=True)
                    start_y += 1
                    for instruct in recipes_instructions:
                        ws.cell(row=start_y, column=start_x).value = f"{instruct[2]}. {instruct[1]}"
                        ws.cell(row=start_y, column=start_x).alignment = op.styles.Alignment(wrap_text=True,
                                                                                             vertical='top')
                        ws.cell(row=start_y, column=start_x).border = border
                        ws.cell(row=start_y, column=start_x).fill = background_fill
                        ws.row_dimensions[start_y].auto_size = True
                        start_y += 1

                    start_x += 3

            workbook.remove_sheet(workbook["Sheet"])
            workbook.save(file_path)
            workbook.close()
    except Exception as e:
        print(f"ERROR IN GENERATING PDFS HAPPENED {datetime.now()}")
        raise e

    return sender_files
