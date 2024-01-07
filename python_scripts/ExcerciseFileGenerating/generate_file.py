import pyodbc
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv
import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill
import subprocess
import glob


load_dotenv("../.env")

server = os.getenv("server")
database = os.getenv("database")
conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes;'
border = Border(left=Side(style='thin', color='6D9712'),
                right=Side(style='thin', color='6D9712'),
                top=Side(style='thin', color='6D9712'),
                bottom=Side(style='thin', color='6D9712'))
background_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')


def generate_file_exercises():
    files = glob.glob('./xlsx_files/*')
    for f in files:
        os.remove(f)
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    sender_files = {}
    try:
        # TrainingForms
        query = f"SELECT * FROM TrainingForms WHERE GenerateFile = 1"
        cursor.execute(query)
        training_forms = cursor.fetchall()

        for training_form in training_forms:
            query_update = "UPDATE TrainingForms SET GenerateFile = 0 WHERE Id = ?"
            cursor.execute(query_update, training_form[0])
            conn.commit()

        for training_form in training_forms:
            query = f"SELECT * FROM Users WHERE Id = ? "
            cursor.execute(query, training_form[1])
            users = cursor.fetchall()
            file_path = f'./xlsx_files/{training_form[0]}_recipes.xlsx'
            sender_files[users[0][3]] = file_path

            query = f"SELECT * FROM Excercises WHERE TrainingFormId = ? "
            cursor.execute(query, training_form[0])
            exercises = cursor.fetchall()

            workbook = op.Workbook()
            days_of_week = {
                'Monday': 0,
                'Tuesday': 1,
                'Wednesday': 2,
                'Thursday': 3,
                'Friday': 4,
                'Saturday': 5,
                'Sunday': 6
            }
            days = training_form[3].split("////")
            for key in days:
                workbook.create_sheet(title=key)
                start_x = 2
                ws = workbook[key]
                ws.sheet_view.showGridLines = False
                exercises_arr_daily = []
                for item in exercises:
                    if item[9] == days_of_week[key]:
                        exercises_arr_daily.append(item)

                for item in exercises_arr_daily:
                    start_y = 2
                    ws.column_dimensions[get_column_letter(start_x)].width = 60
                    ws.cell(row=start_y, column=start_x).value = f"{item[1]} -- {item[3]}"
                    ws.cell(row=start_y, column=start_x).font = op.styles.Font(bold=True)
                    start_y += 1
                    ws.cell(row=start_y, column=start_x).value = f"{item[7].replace('_', ' ').upper()}"
                    start_y += 1
                    ws.cell(row=start_y, column=start_x).value = f"{item[2]}"
                    ws.cell(row=start_y, column=start_x).alignment = op.styles.Alignment(wrap_text=True,
                                                                                         vertical='top')
                    ws.cell(row=start_y, column=start_x).border = border
                    ws.cell(row=start_y, column=start_x).fill = background_fill
                    ws.row_dimensions[start_y].auto_size = True
                    start_y += 2
                    ws.cell(row=start_y, column=start_x).value = f"Coaches comments:"
                    start_y += 1
                    if item[4]:
                        ws.cell(row=start_y, column=start_x).value = f"{item[4]}"
                    else:
                        ws.cell(row=start_y, column=start_x).value = f"No comments yet"
                    ws.cell(row=start_y, column=start_x).alignment = op.styles.Alignment(wrap_text=True,
                                                                                         vertical='top')
                    ws.cell(row=start_y, column=start_x).border = border
                    ws.cell(row=start_y, column=start_x).fill = background_fill
                    ws.row_dimensions[start_y].auto_size = True
                    start_x += 3

            workbook.remove_sheet(workbook["Sheet"])
            workbook.save(file_path)
            workbook.close()
    except Exception as e:
        print(f"ERROR IN GENERATING PDFS HAPPENED {datetime.now()}")
        raise e

    return sender_files

